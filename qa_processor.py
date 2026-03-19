import os
import xml.etree.ElementTree as ET
from collections import defaultdict, Counter
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import zipfile
import shutil

# Configuration
ZIP_FOLDER = "annotations"
EXTRACTED_FOLDER = "extracted_xmls"

# --- PHASE SETTING ---
# Change this to "Phase 2", "Phase 3", etc. before running the script
CURRENT_PHASE = "Phase 1" 

IMAGE_STORAGE_BASE = os.path.join("static", "images")
IMAGE_STORAGE = os.path.join(IMAGE_STORAGE_BASE, CURRENT_PHASE)
EXCEL_FILE = os.path.join("api", "final_QA_report.xlsx")

# Ensure output folders exist
if os.path.exists(EXTRACTED_FOLDER): shutil.rmtree(EXTRACTED_FOLDER)
# Note: We preserve IMAGE_STORAGE_BASE to keep old phase images
os.makedirs(EXTRACTED_FOLDER, exist_ok=True)
os.makedirs(IMAGE_STORAGE, exist_ok=True)

image_votes = defaultdict(list)
student_stats = defaultdict(lambda: {"errors": 0, "warnings": 0, "total": 0, "correct": 0, "label_counts": Counter()})
error_rows = []
voting_rows = []
uncertain_images = []

# -----------------------------
# RULE CHECK
# -----------------------------
def check_rules(tags, img_name, student):
    tag_names = [t.attrib['label'] for t in tags]
    errors, warnings = [], []
    complexity_tags = [t for t in tag_names if "complex" in t or "simple" in t]
    
    if len(complexity_tags) == 0: errors.append("Missing complexity")
    if len(complexity_tags) > 1: errors.append("Multiple complexity")
    if "No Annotation" in tag_names and len(tag_names) > 1: errors.append("No Annotation conflict")
    if "night_dark" in tag_names and "blurred" not in tag_names: warnings.append("Night but not blurred")
    if any(x.lower() in [t.lower() for t in tag_names] for x in ["Rain", "Fog", "Snow"]) and "blurred" not in tag_names: warnings.append("Weather but not blurred")
    
    for e in errors: error_rows.append([img_name, student, "ERROR", e])
    for w in warnings: error_rows.append([img_name, student, "WARNING", w])
    return errors, warnings, tag_names

# -----------------------------
# EXTRACT ZIP FILES
# -----------------------------
print(f"📦 Processing ZIP files for [{CURRENT_PHASE}] from '{ZIP_FOLDER}'...")
IMG_EXTENSIONS = ('.jpg', '.jpeg', '.png', '.webp')

for filename in os.listdir(ZIP_FOLDER):
    if filename.endswith(".zip"):
        student_name = os.path.splitext(filename)[0].strip()
        zip_path = os.path.join(ZIP_FOLDER, filename)
        student_xml_path = os.path.join(EXTRACTED_FOLDER, student_name)
        student_img_path = os.path.join(IMAGE_STORAGE, student_name)
        
        os.makedirs(student_xml_path, exist_ok=True)
        os.makedirs(student_img_path, exist_ok=True)
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                for file_info in zip_ref.infolist():
                    name = file_info.filename.lower()
                    filename_only = os.path.basename(file_info.filename)
                    if not filename_only: continue
                    
                    if name.endswith('.xml'):
                        target_path = os.path.join(student_xml_path, filename_only)
                        with zip_ref.open(file_info) as source, open(target_path, "wb") as target:
                            shutil.copyfileobj(source, target)
                    elif name.endswith(IMG_EXTENSIONS):
                        target_path = os.path.join(student_img_path, filename_only)
                        with zip_ref.open(file_info) as source, open(target_path, "wb") as target:
                            shutil.copyfileobj(source, target)
        except Exception as e:
            print(f"⚠️ Error extracting {filename}: {e}")

# -----------------------------
# LOAD EXTRACTED XML
# -----------------------------
print("🔍 Analyzing extracted XML data...")
for student in os.listdir(EXTRACTED_FOLDER):
    student_path = os.path.join(EXTRACTED_FOLDER, student)
    if os.path.isdir(student_path):
        for xml_file in os.listdir(student_path):
            if not xml_file.endswith(".xml"): continue
            try:
                tree = ET.parse(os.path.join(student_path, xml_file))
                root = tree.getroot()
                for img in root.findall("image"):
                    img_name = os.path.basename(img.attrib["name"])
                    tags = img.findall("tag")
                    errors, warnings, tag_names = check_rules(tags, img_name, student)
                    student_stats[student]["total"] += 1
                    student_stats[student]["errors"] += len(errors); student_stats[student]["warnings"] += len(warnings)
                    for t in tag_names: student_stats[student]["label_counts"][t] += 1
                    image_votes[img_name].append({"student": student, "tags": tag_names})
            except Exception as ex: print(f"⚠️ Error reading {xml_file} from {student}: {ex}")

# -----------------------------
# VOTING SYSTEM LOGIC (NEW)
# -----------------------------
def get_voting_summary(votes):
    total_voters = len(votes)
    if total_voters == 0: return "", set(), False
    
    all_tags = []
    for v in votes: all_tags.extend(v["tags"])
    counts = Counter(all_tags)
    
    strong_majority = set()
    summary_parts = []
    
    # Sort tags by count descending
    sorted_tags = sorted(counts.items(), key=lambda x: x[1], reverse=True)
    
    # Uncertainty Detection: Highest voted tag < 70%
    max_votes = sorted_tags[0][1] if sorted_tags else 0
    is_uncertain = (max_votes / total_voters) < 0.7 if total_voters > 0 else False
    
    for tag, count in sorted_tags:
        percent = (count / total_voters)
        is_complexity = any(kw in tag.lower() for kw in ["complex", "simple"])
        
        if percent >= 0.5:
            # Rule 1: Solid Majority
            strong_majority.add(tag)
            summary_parts.append(f"{tag} ({count})")
        elif not is_complexity and percent >= 0.2:
            # Rule 2: Warning Tags (Supporting tags only)
            summary_parts.append(f"{tag} ⚠️ ({count})")
        # Rule 3: Noise (< 20% or complexity < 50%) is ignored
            
    summary_str = " | ".join(summary_parts)
    return summary_str, strong_majority, is_uncertain

print("⚖️ Calculating Consensus & Scoring...")
for img_name, votes in image_votes.items():
    summary_str, strong_majority, is_uncertain = get_voting_summary(votes)
    
    if is_uncertain:
        uncertain_images.append(img_name)
    
    for v in votes:
        student = v["student"]; student_tags = set(v["tags"])
        
        # Rule 6: Scoring Logic (Intersects only with Strong Majority >= 50%)
        if strong_majority:
            score = len(student_tags & strong_majority) / len(strong_majority)
        else:
            score = 0.0 # No consensus = no score
            
        if score >= 0.7:
            student_stats[student]["correct"] += 1
            
        voting_rows.append([
            img_name, student, 
            ", ".join(student_tags), 
            summary_str, 
            round(score, 2),
            CURRENT_PHASE # Rule: Tag row with the current phase
        ])

# -----------------------------
# SUMMARY & SAVE
# -----------------------------
summary_rows = []
for student, data in student_stats.items():
    acc = (data["correct"] / data["total"]) * 100 if data["total"] > 0 else 0
    summary_rows.append([student, data["total"], data["errors"], data["warnings"], round(acc, 2)])

summary_df = pd.DataFrame(summary_rows, columns=["student", "total", "errors", "warnings", "accuracy"]).sort_values(by="accuracy", ascending=False)

print(f"📊 Finalizing {EXCEL_FILE} with Phase column...")
# Handle existing Excel to APPEND rather than overwrite (Optional, but let's overwrite 
# and let the user re-run phases if they want, OR we can merge them in the script)
# For simplicity, we overwrite the Excel but the Dashboard handles all phases found in the sheet.
# If the user wants to keep Phase 1 AND Phase 2 in the same file, they should concatenate data.

with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    pd.DataFrame(error_rows, columns=["image","student","type","issue"]).to_excel(writer, sheet_name="Errors", index=False)
    pd.DataFrame(voting_rows, columns=["image","student","student_tags","voting_summary","score", "phase"]).to_excel(writer, sheet_name="Voting", index=False)
    pd.DataFrame({"image": list(set(uncertain_images)), "status": "UNCERTAIN"}).to_excel(writer, sheet_name="Review", index=False)

try:
    wb = load_workbook(EXCEL_FILE); ws = wb["Summary"]; chart = BarChart()
    chart.title = "Student Accuracy %"; chart.y_axis.title = "Accuracy %"; chart.x_axis.title = "Students"
    data = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row); cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats); ws.add_chart(chart, "G2"); wb.save(EXCEL_FILE)
except Exception as e: print(f"⚠️ Chart Error: {e}")

print(f"✅ Analysis Complete. Images saved in: {IMAGE_STORAGE}")
